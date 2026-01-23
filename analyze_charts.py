#!/usr/bin/env python3
"""
Analyze Excel files to identify all charts/graphs in each sheet.
"""

import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, DoughnutChart, AreaChart, 
    ScatterChart, RadarChart, BubbleChart, StockChart
)


def get_chart_type_name(chart):
    """Get human-readable chart type name."""
    type_map = {
        'BarChart': 'Bar Chart',
        'LineChart': 'Line Chart', 
        'PieChart': 'Pie Chart',
        'DoughnutChart': 'Doughnut Chart',
        'AreaChart': 'Area Chart',
        'ScatterChart': 'Scatter Chart',
        'RadarChart': 'Radar Chart',
        'BubbleChart': 'Bubble Chart',
        'StockChart': 'Stock Chart',
    }
    class_name = chart.__class__.__name__
    return type_map.get(class_name, class_name)


def get_chart_details(chart):
    """Extract detailed information about a chart."""
    details = {
        'type': get_chart_type_name(chart),
        'class': chart.__class__.__name__,
        'title': str(chart.title) if chart.title else None,
        'series_count': len(chart.series) if hasattr(chart, 'series') else 0,
        'series': [],
    }
    
    # Get bar chart specific info
    if hasattr(chart, 'type'):
        details['subtype'] = chart.type
    if hasattr(chart, 'grouping'):
        details['grouping'] = chart.grouping
    if hasattr(chart, 'barDir'):
        details['bar_direction'] = chart.barDir
        
    # Get series information
    if hasattr(chart, 'series'):
        for i, series in enumerate(chart.series):
            series_info = {'index': i}
            if hasattr(series, 'title') and series.title:
                series_info['title'] = str(series.title)
            if hasattr(series, 'val') and series.val:
                series_info['values_ref'] = str(series.val.numRef.f) if hasattr(series.val, 'numRef') and series.val.numRef else None
            if hasattr(series, 'cat') and series.cat:
                series_info['categories_ref'] = str(series.cat.numRef.f) if hasattr(series.cat, 'numRef') and series.cat.numRef else None
            details['series'].append(series_info)
    
    return details


def analyze_excel_file(file_path):
    """Analyze an Excel file and return information about all charts."""
    print(f"\n{'='*80}")
    print(f"Analyzing: {file_path.name}")
    print('='*80)
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  ERROR loading file: {e}")
        return None
    
    file_info = {
        'file_name': file_path.name,
        'file_path': str(file_path),
        'sheets': []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            'sheet_name': sheet_name,
            'charts': [],
            'data_range': f"A1:{ws.dimensions.split(':')[1] if ':' in (ws.dimensions or 'A1:A1') else 'A1'}"
        }
        
        print(f"\n  Sheet: '{sheet_name}'")
        print(f"    Data range: {sheet_info['data_range']}")
        
        # Get charts from the sheet
        if hasattr(ws, '_charts') and ws._charts:
            charts = ws._charts
        else:
            charts = []
        
        if not charts:
            print(f"    Charts: None")
        else:
            print(f"    Charts found: {len(charts)}")
            
            for idx, chart in enumerate(charts):
                chart_info = get_chart_details(chart)
                chart_info['index'] = idx
                sheet_info['charts'].append(chart_info)
                
                print(f"\n      Chart {idx + 1}:")
                print(f"        Type: {chart_info['type']}")
                if chart_info.get('subtype'):
                    print(f"        Subtype: {chart_info['subtype']}")
                if chart_info.get('grouping'):
                    print(f"        Grouping: {chart_info['grouping']}")
                if chart_info.get('bar_direction'):
                    print(f"        Direction: {chart_info['bar_direction']}")
                print(f"        Title: {chart_info['title']}")
                print(f"        Series count: {chart_info['series_count']}")
                
                for series in chart_info['series']:
                    print(f"          Series {series['index']}: {series.get('title', 'Untitled')}")
                    if series.get('values_ref'):
                        print(f"            Values: {series['values_ref']}")
        
        file_info['sheets'].append(sheet_info)
    
    wb.close()
    return file_info


def analyze_data_structure(file_path):
    """Analyze the data structure of an Excel file to understand potential chart data."""
    print(f"\n  Analyzing data structure...")
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        return None
        
    data_info = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers (first row)
        headers = []
        for col in range(1, min(ws.max_column + 1, 50)):  # Limit to first 50 columns
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
        
        # Sample first few rows
        sample_data = []
        for row in range(2, min(6, ws.max_row + 1)):
            row_data = {}
            for col, header in enumerate(headers, start=1):
                cell_value = ws.cell(row=row, column=col).value
                row_data[header] = str(cell_value) if cell_value else None
            sample_data.append(row_data)
        
        data_info[sheet_name] = {
            'headers': headers,
            'row_count': ws.max_row,
            'sample_data': sample_data
        }
        
        print(f"\n    Sheet '{sheet_name}':")
        print(f"      Headers: {headers[:10]}{'...' if len(headers) > 10 else ''}")
        print(f"      Row count: {ws.max_row}")
    
    wb.close()
    return data_info


def main():
    script_dir = Path(__file__).parent
    
    # Files to analyze
    excel_files = [
        script_dir / "10 Oct25 SAFETY and QC COST REPORT - old.xlsx",
        script_dir / "Copy of 10 Oct25 Divisional OH report (002).xlsx",
        script_dir / "Copy of Oct25 Equipment Results.xlsx",
        script_dir / "Oct fuel spend trend.xlsx",
    ]
    
    input_file = script_dir / "input" / "YTD Indirect-G&A Cost.xlsx"
    
    all_results = {
        'source_files': [],
        'input_file': None
    }
    
    # Analyze each source file
    for file_path in excel_files:
        if file_path.exists():
            file_info = analyze_excel_file(file_path)
            if file_info:
                data_info = analyze_data_structure(file_path)
                file_info['data_structure'] = data_info
                all_results['source_files'].append(file_info)
        else:
            print(f"\nFile not found: {file_path}")
    
    # Analyze input file structure
    print(f"\n{'='*80}")
    print("INPUT FILE ANALYSIS")
    print('='*80)
    
    if input_file.exists():
        input_info = analyze_excel_file(input_file)
        if input_info:
            input_data = analyze_data_structure(input_file)
            input_info['data_structure'] = input_data
            all_results['input_file'] = input_info
    
    # Save results to JSON
    output_path = script_dir / "chart_analysis_results.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, default=str)
    
    print(f"\n\nResults saved to: {output_path}")
    
    # Summary
    print(f"\n{'='*80}")
    print("SUMMARY")
    print('='*80)
    
    total_charts = 0
    for file_info in all_results['source_files']:
        file_charts = sum(len(sheet['charts']) for sheet in file_info['sheets'])
        total_charts += file_charts
        print(f"\n{file_info['file_name']}:")
        for sheet in file_info['sheets']:
            if sheet['charts']:
                print(f"  {sheet['sheet_name']}: {len(sheet['charts'])} chart(s)")
                for chart in sheet['charts']:
                    print(f"    - {chart['type']}: {chart['title']}")
    
    print(f"\nTotal charts found: {total_charts}")
    
    return all_results


if __name__ == "__main__":
    main()




