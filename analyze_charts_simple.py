#!/usr/bin/env python3
"""
Simple chart analysis - extract chart titles and types from Excel files.
"""

import json
from pathlib import Path
from openpyxl import load_workbook


def extract_title_text(title_obj):
    """Extract text from a chart title object."""
    if title_obj is None:
        return None
    
    # Try to get the title as string first
    try:
        if hasattr(title_obj, 'tx') and title_obj.tx:
            if hasattr(title_obj.tx, 'rich') and title_obj.tx.rich:
                rich = title_obj.tx.rich
                if hasattr(rich, 'p'):
                    for para in rich.p:
                        if hasattr(para, 'r'):
                            for run in para.r:
                                if hasattr(run, 't') and run.t:
                                    return run.t
    except Exception:
        pass
    
    return str(title_obj) if title_obj else None


def get_chart_type(chart):
    """Get chart type name."""
    class_name = chart.__class__.__name__
    
    # Get more detail
    subtype = ""
    if hasattr(chart, 'type'):
        subtype = f" ({chart.type})"
    if hasattr(chart, 'barDir'):
        subtype = f" ({chart.barDir})"
    if hasattr(chart, 'grouping'):
        subtype += f" [{chart.grouping}]"
        
    return class_name + subtype


def analyze_excel_charts(file_path):
    """Analyze charts in an Excel file."""
    print(f"\n{'='*70}")
    print(f"FILE: {file_path.name}")
    print('='*70)
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        return []
    
    all_charts = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        charts = ws._charts if hasattr(ws, '_charts') else []
        
        if charts:
            print(f"\n  Sheet: '{sheet_name}' - {len(charts)} chart(s)")
            
            for idx, chart in enumerate(charts):
                title = extract_title_text(chart.title)
                chart_type = get_chart_type(chart)
                series_count = len(chart.series) if hasattr(chart, 'series') else 0
                
                # Get series names
                series_names = []
                if hasattr(chart, 'series'):
                    for s in chart.series:
                        if hasattr(s, 'tx') and s.tx:
                            try:
                                if hasattr(s.tx, 'strRef') and s.tx.strRef:
                                    series_names.append(str(s.tx.strRef.f))
                                elif hasattr(s.tx, 'v'):
                                    series_names.append(str(s.tx.v))
                            except:
                                pass
                
                chart_info = {
                    'file': file_path.name,
                    'sheet': sheet_name,
                    'index': idx,
                    'type': chart_type,
                    'title': title,
                    'series_count': series_count,
                    'series_names': series_names
                }
                all_charts.append(chart_info)
                
                print(f"    [{idx+1}] {chart_type}")
                print(f"        Title: {title or '(no title)'}")
                print(f"        Series: {series_count}")
                if series_names:
                    print(f"        Series names: {series_names[:3]}{'...' if len(series_names) > 3 else ''}")
        else:
            print(f"\n  Sheet: '{sheet_name}' - No charts")
    
    wb.close()
    return all_charts


def analyze_data_headers(file_path):
    """Get data headers from Excel sheets."""
    print(f"\n  Data Structure:")
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        return {}
    
    data_info = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers (first row)
        headers = []
        for col in range(1, min(ws.max_column + 1, 20)):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value)[:40])
        
        if headers:
            data_info[sheet_name] = {
                'headers': headers,
                'row_count': ws.max_row
            }
            print(f"    {sheet_name}: {ws.max_row} rows, headers: {headers[:5]}{'...' if len(headers) > 5 else ''}")
    
    wb.close()
    return data_info


def main():
    script_dir = Path(__file__).parent
    
    # Files to analyze
    files = [
        script_dir / "10 Oct25 SAFETY and QC COST REPORT - old.xlsx",
        script_dir / "Copy of 10 Oct25 Divisional OH report (002).xlsx",
        script_dir / "Copy of Oct25 Equipment Results.xlsx",
        script_dir / "Oct fuel spend trend.xlsx",
    ]
    
    input_file = script_dir / "input" / "YTD Indirect-G&A Cost.xlsx"
    
    all_charts = []
    all_data = {}
    
    print("\n" + "="*70)
    print("CHART ANALYSIS REPORT")
    print("="*70)
    
    # Analyze each file
    for file_path in files:
        if file_path.exists():
            charts = analyze_excel_charts(file_path)
            all_charts.extend(charts)
            data = analyze_data_headers(file_path)
            all_data[file_path.name] = data
        else:
            print(f"\nFile not found: {file_path.name}")
    
    # Analyze input file
    print("\n" + "="*70)
    print("INPUT FILE STRUCTURE")
    print("="*70)
    
    if input_file.exists():
        input_data = analyze_data_headers(input_file)
        all_data['INPUT'] = input_data
    
    # Summary
    print("\n" + "="*70)
    print("SUMMARY OF ALL CHARTS")
    print("="*70)
    
    # Group by file
    by_file = {}
    for chart in all_charts:
        fname = chart['file']
        if fname not in by_file:
            by_file[fname] = []
        by_file[fname].append(chart)
    
    for fname, charts in by_file.items():
        print(f"\n{fname}:")
        for c in charts:
            print(f"  - {c['type']}: {c['title'] or '(untitled)'} ({c['series_count']} series)")
    
    print(f"\n\nTOTAL CHARTS: {len(all_charts)}")
    
    # Chart types breakdown
    types = {}
    for c in all_charts:
        t = c['type'].split(' ')[0]  # Get main type
        types[t] = types.get(t, 0) + 1
    
    print("\nBy Type:")
    for t, count in sorted(types.items(), key=lambda x: -x[1]):
        print(f"  {t}: {count}")
    
    # Save summary
    summary = {
        'charts': all_charts,
        'data_structure': all_data
    }
    
    with open(script_dir / 'chart_summary.json', 'w') as f:
        json.dump(summary, f, indent=2)
    
    print(f"\nSummary saved to: chart_summary.json")
    
    return all_charts


if __name__ == "__main__":
    main()




